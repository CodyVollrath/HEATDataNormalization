from openpyxl import Workbook,load_workbook

"""
Class WorkBookReader is responsible for abstracting the API openpyxl
:Author Cody Vollrath
:Company Southwire
:Version 1.0
"""
class WorkBookReader:
    """
    Constructor: Creates an instance of the WorkBookReader class
    :precondition workbook must not be None and contain .xlsx within the name
    :postcondition an instance of the WorkbookReader will be created

    """
    def __init__(self, workbook):
        if workbook is None:
            raise ValueError("Workbook name must not be 'None'")
        if ".xlsx" not in workbook.lower():
            raise ValueError("Workbook is not a compatible file")
        self.wb2 = load_workbook(workbook)
    """
    Simply returns the workbook set by the user in the constructor of WorkBookReader
    :return workbook object
    """
    def getWorkbookObject(self):
        return self.wb2
    """
    Gets the data from the column or row depending on the user input for the colOrRow parameter
    if colOrRow is entered as a numeric string like '1' then row 1 will be read,
    likewise, if 'A' is entered, then column 'A' will be read
    
    :precondition colOrRow is not None && worksheet is not None
    :postcondition none
    :return data in the column or row specified by the user
    """
    def getColumnOrRowOfWorkSheet(self, worksheet, colOrRow):
        if worksheet is None:
            raise ValueError("Worksheet must not be 'None'")
        if colOrRow is None:
            raise ValueError("ColOrRow must not be 'None'")
        itemList = []
        try:
            ws = self.wb2[worksheet]

            for cell in ws[colOrRow]:
                itemList.append(cell.value)
        except:
            print("Errors in processing - Check that " + worksheet + " is a valid sheet name in the workbook")
        return itemList

    """
    Gets the value of a cell entered by the user.
    :precondition cell is not None && worksheet is not None
    :postcondition none
    :return cell value for the cell specfied
    """
    def getValueOfCell(self, worksheet, cell):
        if worksheet is None:
            raise ValueError("Worksheet must not be 'None'")
        if cell is None:
            raise ValueError("cell must not be 'None'")
        return self.wb2[worksheet][cell].value